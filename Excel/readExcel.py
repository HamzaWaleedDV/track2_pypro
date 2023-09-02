import openpyxl
from pathlib import Path

excelFile = openpyxl.load_workbook(Path.home() / Path("OneDrive", "Desktop", "test.xlsx"))
print(excelFile.sheetnames)

Sheet1 = excelFile['Sheet1']
print(Sheet1.title)

activesheet = excelFile.active
print(activesheet.title)

print(Sheet1["A1"].value)
print(Sheet1["B1"].value)
print(Sheet1["C1"].value)
print(Sheet1['C1'].row)
print(Sheet1['C1'].column)
print(Sheet1['C1'].coordinate)

print(Sheet1.cell(row=1, column=2).value)


for n in range(1, Sheet1.max_row+1):
    print(n, Sheet1[f"B{n}"].value)


print('-' * 50)

result = 0

for n in range(1, 7):
    print('The name is:', Sheet1[f"A{n}"].value, ', The salary is:', Sheet1[f'B{n}'].value)
    result += Sheet1[f'B{n}'].value
print(f'All salaries is: {result}')

print('-' * 50)
print(Sheet1.max_row)
print(Sheet1.max_column)