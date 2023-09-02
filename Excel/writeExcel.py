import openpyxl
from pathlib import Path

excelFile = openpyxl.Workbook()
print(excelFile.sheetnames)

#change sheet name
excelSheet = excelFile.active
excelSheet.title = 'Hamza'

#create sheets
excelFile.create_sheet()
excelFile.create_sheet()
excelFile.create_sheet(index=1, title='Waleed')
excelFile.create_sheet(index=2, title='Mahmoud')


#delete sheet
del excelFile['Sheet1']

#Write to sheet
sheet = excelFile['Waleed']
sheet['A1'] = 'Hello world'
print(sheet['A1'].value)

#Exam
sheet1 = excelFile['Hamza']
sheet1['C1'] = 'Hamza'
sheet1['C3'] = 'Waleed'
sheet1['C2'] = 'Mahmoud'

#OR
names = ['Hamza', 'Waleed', 'Mahmoud', 'Nasr']

for number, name in enumerate(names, 1):
    sheet1 = excelFile['Hamza']
    sheet1[f'C{number}'] = name

#save excel file
excelFile.save(filename = Path.home() / Path("OneDrive", "Desktop", "newExcel.xlsx"))